import os
import re
import openpyxl
import pandas as pd


def get_date(page_name: str) -> str:
    return re.search(r'(\d{2}-\d{2}-\d{2})', page_name).group(0)


def get_frequency(page_name: str) -> str:
    return re.search(r'(\d+kHz)', page_name).group(0)


def get_exposure_time(page_name: str) -> str:
    pattern_string = re.search(r'(\s+\d+[\s\'])', page_name).group(0).strip()
    return pattern_string if "'" not in pattern_string else pattern_string[:len(pattern_string) - 1]


def get_supply_delay(page_name: str) -> str:
    return re.search(r'(\d+[dw])', page_name).group(0)


def get_method_type(file_name: str) -> str:
    return "DBD" if "DBD" in file_name else "Torch"


def print_metadata_list(data: list, quantity: int = 1) -> None:
    quantity = quantity if quantity < len(data) else len(data) - 1
    print(f'Printing the first {quantity} entries out of {len(data)}\n')

    for data_dict in data[:quantity]:
        print(f'date: {data_dict["date"]}')
        print(f'type: {data_dict["type"]}')
        print(f'frequency: {data_dict["frequency"]}')
        print(f'exposure time: {data_dict["exposure time"]}')
        print(f'supply delay: {data_dict["supply delay"]}\n')


def get_database(path: str) -> list:
    xlsx_files = [path + "/" + file for file in os.listdir(path) if file.endswith(".xlsx")]
    series_list = list()

    for path in xlsx_files:
        wb_obj = openpyxl.load_workbook(path)

        for page in wb_obj.sheetnames:
            sheet = wb_obj[page]

            for column in sheet.iter_cols(values_only=True):
                time_data = pd.Series(data=column)

                if time_data is None or time_data.empty or time_data.isna().sum() == time_data.size:
                    continue

                # reject series of the type (15' RT) or 3Khz + 3Khz
                try:
                    data_dict = {
                        "date": get_date(page),
                        "type": get_method_type(path),
                        "frequency": get_frequency(page),
                        "exposure time": get_exposure_time(page),
                        "supply delay": get_supply_delay(page),
                        "time_data": time_data
                    }

                    series_list.append(data_dict)

                except AttributeError:
                    pass

    return series_list
