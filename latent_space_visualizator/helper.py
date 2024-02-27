import os
import pickle
import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from matplotlib.figure import Figure
from scipy.interpolate import interp1d
import models

names = {
    "exposition time": "exposure time",
    "DBD vs PT": "type",
    "storage time": "supply delay",
    "power or freq": "frequency",
}


def atmf(data: np.array, win_size: int, alpha: int) -> list:
    n = len(data)
    result = [0] * (n - win_size)

    assert win_size > alpha > 0 and alpha % 4 == 0 and win_size % 2 == 0

    for i in range(win_size // 2, n - win_size // 2):
        window = data[i - win_size // 2:i + win_size // 2]
        window.sort()
        result[i - win_size // 2] = sum(window[alpha // 4:win_size - (alpha // 4) * 3]) / (win_size - alpha)

    return [result[0]] * (win_size // 2) + result + [result[-1]] * (win_size // 2)


def load_autoencoder(dim: int) -> models.FullAutoencoder:
    full_autoencoder = models.FullAutoencoder(models.Autoencoder(), dim)
    full_autoencoder.build(input_shape=(None, 1800, 1))
    full_autoencoder.load_weights(f"./weights/FullCOnvAE{str(dim)}Dbis.h5")
    return full_autoencoder


def load_database(path: str) -> tuple[dict, np.array, np.array]:
    with open(path, 'rb') as f:
        database = pickle.load(f)

    all_np_array = np.array([data_dict["time_data"].to_numpy() for data_dict in database])

    all_smoothed = np.array([atmf(x.tolist(), 80, 40) for x in all_np_array])
    all_smoothed = all_smoothed[:, :, np.newaxis]
    return database, all_np_array, all_smoothed


def setup_frame(title: str) -> tuple[Figure, list]:
    fig, ax = plt.subplots(nrows=1, ncols=2, figsize=(14, 6))
    plt.subplots_adjust(left=0.3)
    fig.canvas.manager.set_window_title(title)
    fig.suptitle('Latent Space Visualization', fontsize=16)
    ax[0].set(xlabel='first dimension', ylabel='second dimension')
    ax[1].set(xlabel='third dimension', ylabel='fourth dimension')
    return fig, ax


def create_colors(label: str, database: dict, ax: list) -> tuple[dict, list]:
    unique_values = {i[label] for i in database}

    color_palette = plt.cm.get_cmap('tab10', len(unique_values))
    colors_dict = {element: color_palette(i) for (i, element) in enumerate(unique_values)}

    if "0d" in colors_dict.keys() and "100s" in colors_dict.keys():
        colors_dict["0d"] = colors_dict["100s"]

    legend_handles = [plt.scatter([], [], color=colors_dict[i], label=i) for i in colors_dict]

    ax[0].legend(handles=legend_handles)
    ax[1].legend(handles=legend_handles)

    return colors_dict, legend_handles


def get_colors_list(database: dict, colors: dict, label: str) -> list:
    return [colors[i[label]] for i in database]


def scale(ax, values):
    x_min, x_max = min(values[0]), max(values[0])
    y_min, y_max = min(values[1]), max(values[1])

    ax.set_xlim(x_min - 0.1 * (x_max - x_min), x_max + 0.1 * (x_max - x_min))
    ax.set_ylim(y_min - 0.1 * (y_max - y_min), y_max + 0.1 * (y_max - y_min))


def parser(path: str) -> np.array:
    xlsx_files = [path + "/" + file for file in os.listdir(path) if file.endswith(".xlsx")]
    series_list = list()

    for path in xlsx_files:
        wb_obj = openpyxl.load_workbook(path)

        for page in wb_obj.sheetnames:
            sheet = wb_obj[page]

            for column in sheet.iter_cols(values_only=True):

                if column is None or len(column) == 0:
                    continue

                time_data = np.array(column)

                series_list.append(time_data)

    return _interpolate_series(series_list)


def _interpolate_series(series_list: list) -> np.array:
    interpolate_list = list()

    x_interpolated = np.linspace(0, len(series_list[0]) - 1, 1800)
    x_original = np.arange(len(series_list[0]))

    for data in series_list:
        interpolated_function = interp1d(x_original, data, kind='linear')
        interpolate_list.append(interpolated_function(x_interpolated))

    return np.array(interpolate_list)
