import os
import pickle
import re
import openpyxl
import pandas as pd
import numpy as np
from customLib.filters import atmf


def get_exposure_time(page_name: str) -> str:
    pattern_string = re.search(r'(\s+\d+[\s\'])', page_name).group(0).strip()
    return pattern_string if "'" not in pattern_string else pattern_string[:len(pattern_string) - 1]


def print_metadata_list(data: list, quantity: int = 1) -> None:
    quantity = quantity if quantity < len(data) else len(data) - 1
    print(f'Printing the first {quantity} entries out of {len(data)}\n')

    for data_dict in data[:quantity]:
        print(f'date: {data_dict["date"]}')
        print(f'type: {data_dict["type"]}')
        print(f'frequency: {data_dict["frequency"]}')
        print(f'exposure time: {data_dict["exposure time"]}')
        print(f'supply delay: {data_dict["supply delay"]}\n')


def multi_page_dict_creator(page: str, path: str, time_data: pd.Series) -> dict:
    return {
        "date": re.search(r'(\d{2}-\d{2}-\d{2})', page).group(0),
        "type": "DBD" if "DBD" in path else "DII",
        "frequency": re.search(r'(\d+kHz)', page).group(0),
        "exposure time": f"{int(get_exposure_time(page)) * 60}s",
        "supply delay": re.search(r'(\d+[dw])', page).group(0),
        "time_data": time_data
    }


def single_page_dict_creator(page: str, path: str, time_data: pd.Series) -> dict:
    data_list = path[:-5].split('/')[-1].split("_")
    return {
        "date": f"{data_list[1][:2]}-{data_list[1][2:4]}-{data_list[1][4:]}",
        "type": data_list[2],
        "frequency": data_list[4],
        "exposure time": data_list[3],
        "supply delay": data_list[5],
        "time_data": time_data
    }


def parser(path: str, dict_creator=None):
    xlsx_files = [path + "/" + file for file in os.listdir(path) if file.endswith(".xlsx")]
    series_list = list()

    for path in xlsx_files:
        wb_obj = openpyxl.load_workbook(path)

        for page in wb_obj.sheetnames:
            sheet = wb_obj[page]

            for column in sheet.iter_cols(values_only=True):
                time_data = pd.Series(data=column)

                if time_data is None or time_data.empty or time_data.isna().sum() == time_data.size:
                    continue

                try:
                    data_dict = dict_creator(page, path, time_data)
                    series_list.append(data_dict)

                except AttributeError:
                    pass

    return series_list


def create_db(path: str) -> list:
    db = parser(path + "/multi-page", multi_page_dict_creator) + parser(path + "/single-page", single_page_dict_creator)
    no_duplicates_db = list()

    for i in range(len(db)):
        for j in range(i + 1, len(db)):
            if db[i]["time_data"].equals(other=db[j]["time_data"]):
                break
            if j == len(db) - 1 and not db[i]["time_data"].hasnans:
                no_duplicates_db.append(db[i])

    print(f"len db with duplicates: {len(db)}")
    print(f"len db without duplicates: {len(no_duplicates_db)}")

    return no_duplicates_db


def load_database(path: str, show_phy=False) -> tuple[list, np.array, np.array]:
    with open(path, 'rb') as f:
        database = pickle.load(f)

    all_np_array = np.array([data_dict["time_data"].to_numpy() for data_dict in database])

    if show_phy:
        physical_data = parser("data/phy_data")
        all_np_array = np.concatenate((all_np_array, physical_data))
        for i in range(len(physical_data)):
            database.append(
                {"time_data": physical_data[i], "exposure time": "physical", "type": "physical", "supply delay": "physical", "frequency": "physical"})

    all_smoothed = np.array([atmf(x.tolist(), 80, 40) for x in all_np_array])
    all_smoothed = all_smoothed[:, :, np.newaxis]
    all_np_array = all_np_array[:, :, np.newaxis]

    return database, all_np_array, all_smoothed
